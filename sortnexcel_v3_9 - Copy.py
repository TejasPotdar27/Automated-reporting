import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
#from datetime import timedelta
import os
import pdfplumber
import pandas as pd
import openpyxl
import subprocess
import shutil
import send2trash

def extract_sentences_with_keywords(pdf_path, keywords):
    sentences = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            sentences.extend(text.split('\n'))
    filtered_sentences = []
    for sentence in sentences:
        for keyword in keywords:
            if sentence.strip().endswith(keyword):
                filtered_sentence = sentence.strip()[:-len(keyword)].strip()
                filtered_sentences.append(filtered_sentence)
                break
    return filtered_sentences

def extract_duration_from_pdf(pdf_path):
    duration = None
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]
        first_half_text = first_page.extract_text()
        lines = first_half_text.split('\n')
        for line in lines:
            if 'Duration' in line:
                duration = ':'.join(line.split(':')[-3:]).strip()
                break
    return duration

def extract_date_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        for line in first_page_text.split('\n'):
            if line.startswith("Test begin:"):
                date_str = line.split("Test begin:")[1].split(" ")[1]
                try:
                    # Try the CLI format first (YYYY/MM/DD)
                    return datetime.strptime(date_str, "%Y/%m/%d").strftime("%Y-%m-%d")
                except ValueError:
                    try:
                        # Try the manual format second (MM/DD/YYYY)
                        return datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                    except ValueError:
                        return None
    return None
    
def extract_Testtimes_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        start_time, end_time = None, None

        for line in first_page_text.split('\n'):
            if line.startswith("Test begin:"):
                parts = line.split("Test begin:")[1].strip().split(" ", 1)
                if len(parts) > 1:
                    start_time = parts[1]  # this includes time and timezone
            elif line.startswith("Test end:"):
                parts = line.split("Test end:")[1].strip().split(" ", 1)
                if len(parts) > 1:
                    end_time = parts[1]

        return start_time, end_time    
    
def extract_times_from_pdf(pdf_path):
    download_time = None
    activation_time = None
    installation_time = None
    down_time = None
    with pdfplumber.open(pdf_path) as pdf:
        #first_page_text = pdf.pages[0].extract_text()
        pages_to_consider = 0
        
        # Search for "Table of Contents" in the first three pages
        for i in range(min(3, len(pdf.pages))):  # Limit to the first 3 pages
            page_text = pdf.pages[i].extract_text()
            if "Table of Contents" in page_text:
                pages_to_consider = i + 1  # Include the page with "Table of Contents"
                break
        
        # If "Table of Contents" is not found, default to the first 3 pages
        if pages_to_consider == 0:
            pages_to_consider = min(3, len(pdf.pages))
        
        # Process the determined range of pages
        for i in range(pages_to_consider):
            page_text = pdf.pages[i].extract_text()
            
            for line in page_text.split('\n'):
                if "Download Time" in line:
                    #download_time = line.split(":")[1].strip()
                    download_time = ':'.join(line.split(':')[-3:]).strip()
                if "Activation Time" in line:
                    activation_time = ':'.join(line.split(':')[-3:]).strip()
                if "Installation Time" in line:
                    installation_time = ':'.join(line.split(':')[-3:]).strip()
                if "Down Time" in line:
                    down_time = ':'.join(line.split(':')[-3:]).strip()    
                
    return download_time, activation_time, installation_time, down_time

# when Test script objective is single line
#def extract_result_from_detailed_report(pdf_path):
    # with pdfplumber.open(pdf_path) as pdf:
        # first_page_text = pdf.pages[0].extract_text()
        # print ({first_page_text})
        # result_index = first_page_text.find("Function Block [Result]")
        # if result_index != -1:
            # result_status = first_page_text[result_index - 5:result_index].strip()
            # return result_status
    # return None
    
# when Test script objective is multi line
def extract_result_from_detailed_report(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        
        # Define a search area around the word "Test Script Objective"
        result_search_start = first_page_text.find("Test Script Objective")
        
        if result_search_start != -1:
            # Limit search to a few hundred characters following "Test Script Objective"
            text_after_result = first_page_text[result_search_start:result_search_start + 600]
            
            # Check for "Pass" or "Fail" within this range
            if "Fail" in text_after_result:
                return "Fail"
            elif "Pass" in text_after_result:
                return "Pass"

    return "NA"  # Return "NA" if no result is found
    
    
def extract_test_case_result_from_vtreport(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        result_index = first_page_text.find("Test Case Result")
        if result_index != -1:
            test_case_result = first_page_text[result_index - 5:result_index].strip()
            if test_case_result not in ["None", "Pass", "Fail"]:
                test_case_result = "Inconclusive"
            return test_case_result        
    return None
  
# def extract_campaign_details_from_pdf(pdf_path):
    # campaign_name = None
    # campaign_url = None
    
    # with pdfplumber.open(pdf_path) as pdf:
        # for page_number in range(1, len(pdf.pages)):  # Start from the second page (index 1)
            # page_text = pdf.pages[page_number].extract_text()
            # #print(f"Page {page_number + 1} Text: {page_text}")
            # for line in page_text.split('\n'):
                # if line.startswith("Measured value: Campaign Name"):
                    # campaign_name = line.split("=")[1].strip()
                    # #print(f"Extracted Campaign Name: {campaign_name}")
                # if line.startswith("Measured value: Campaign URL") and line.endswith("com"):
                    # campaign_url = line.split("=")[1].strip()
                    # #print(f"Extracted Campaign URL: {campaign_url}")
    
    # return campaign_name, campaign_url
    
def extract_campaign_details_from_pdf(pdf_path):
    campaign_name = ""
    campaign_url = ""
    capture_name = False
    capture_url = False
    
    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) > 0:
            # page_text = pdf.pages[0].extract_text()
            # lines = page_text.split('\n')
            
            pages_to_consider = 0
        
            # Search for "Table of Contents" in the first three pages
            for i in range(min(3, len(pdf.pages))):  # Limit to the first 3 pages
                page_text = pdf.pages[i].extract_text()
                if "Table of Contents" in page_text:
                    pages_to_consider = i + 1  # Include the page with "Table of Contents"
                    break
            
            # If "Table of Contents" is not found, default to the first 3 pages
            if pages_to_consider == 0:
                pages_to_consider = min(3, len(pdf.pages))
            
            # Process the determined range of pages
            for i in range(pages_to_consider):
                page_text = pdf.pages[i].extract_text()
                
                for line in page_text.split('\n'):
                    # Handling Campaign Name splited in 2 lines
                    if line.startswith("Campaign Name"):
                        capture_name = True
                        campaign_name += line.split(":")[1].strip()
                    elif capture_name:
                        if line.startswith("Campaign URL"):
                            capture_name = False
                            capture_url = True
                        else:
                            campaign_name += " " + line.strip()
                    
                    # Handling Campaign URL splited in multiple lines
                    if line.startswith("Campaign URL"):
                        capture_url = True
                        campaign_url += line.split(":")[1].strip()
                    elif capture_url:
                        if line.startswith("Campaign Type"):
                            capture_url = False
                        else:
                            campaign_url += line.strip()
    
    # Clean up the extracted values
    campaign_name = campaign_name.strip()
    campaign_url = campaign_url.replace("\n", "").replace(" ", "").strip()
    
    return campaign_name, campaign_url
    
def extract_campgtyp_prgID_rssi_setup_details_from_pdf(pdf_path):
    
    campaign_type = None
    program_before = ""
    program_after = ""
    rssi = None
    setup = None
    computer_name = None
    capture2_name = False
    capture3_name = False
    
    with pdfplumber.open(pdf_path) as pdf:
        #first_page_text = pdf.pages[0].extract_text()
        #print(first_page_text)
        # Initialize the search range and flag for "Table of Contents"
        pages_to_consider = 0
        
        # Search for "Table of Contents" in the first three pages
        for i in range(min(3, len(pdf.pages))):  # Limit to the first 3 pages
            page_text = pdf.pages[i].extract_text()
            if "Table of Contents" in page_text:
                pages_to_consider = i + 1  # Include the page with "Table of Contents"
                break
        
        # If "Table of Contents" is not found, default to the first 3 pages
        if pages_to_consider == 0:
            pages_to_consider = min(3, len(pdf.pages))
        
        # Process the determined range of pages
        for i in range(pages_to_consider):
            page_text = pdf.pages[i].extract_text()
            
            # Process the text content for relevant details
            for line in page_text.split('\n'):
                #print(f"Processing line: {line}")
                # if "Program ID before Test" in line:
                    # program_before = line.split(":")[1].strip()
                # Handling Program ID before test splited in 2 lines
                if line.startswith("Program ID before Test"):
                    capture2_name = True
                    program_before += line.split(":")[1].strip()
                elif capture2_name:
                    if line.startswith("Download Time"):
                        capture2_name = False
                    else:
                        program_before += " " + line.strip()    
                # if "Program ID after Test" in line:
                    # program_after = line.split(":")[1].strip()
                # Handling Program ID after test splited in 2 lines
                if line.startswith("Program ID after Test"):
                    capture3_name = True
                    program_after += line.split(":")[1].strip()
                elif capture3_name:
                    if line.startswith("Hardware"):
                        capture3_name = False
                    else:
                        program_after += " " + line.strip() 
                        
                if "Campaign Type" in line:
                    campaign_type = line.split(":")[1].strip()    
                if "RSSI Strength" in line:
                    rssi =  line.split(":")[1].strip()
                    #print(f"Extracted RSSI: {rssi}")
                if "Windows Computer Name:" in line:
                    computer_name = line.split(":")[1].strip()
                
     # Clean up the extracted values
    program_before = program_before.replace("\n", "").replace(" ", "").strip()
    program_after = program_after.replace("\n", "").replace(" ", "").strip()
    
    
    setup = determine_setup(computer_name)
    
    return campaign_type, program_before, program_after, rssi, setup
    
def determine_setup(computer_name):
    if computer_name == "12":
        return "1"
    elif computer_name == "12":
        return "2"
    elif computer_name == "PD-73":
        return "1"
    elif computer_name == "PD-74":
        return "2"
    elif computer_name == "PD-77":
        return "3"    
    else:
        return "NA"

def determine_ecu_type(test_pattern_name):
    if "CGW" in test_pattern_name:
        return "2Bank"
    elif "IVH" in test_pattern_name:
        return "1Bank"
    elif "ICM" in test_pattern_name:
        return "2Bank" 
    elif "FI" in test_pattern_name:
        return "1Bank"    
    # Add more conditions here as needed
    else:
        return "NA"
        
def write_sentences_to_excel(data, excel_path):
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 120

    ws.column_dimensions['AF'].width = 100       #Adjust Comments col width

    for col in ws.columns:
        if col[0].column != 6:
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].auto_size = True

    wb.save(excel_path)

def browse_input_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        input_entry.delete(0, tk.END)
        input_entry.insert(0, folder_path)

def browse_output_dir():
    output_dir = filedialog.askdirectory()
    if output_dir:
        output_entry.delete(0, tk.END)
        output_entry.insert(0, output_dir)

def generate_excel():
    input_dir = input_entry.get()
    output_dir = output_entry.get()
    keyword = keyword_entry.get()

    if not input_dir or not output_dir :
        result_label.config(text="Please select input folder and output directory.")
        return

    if not keyword :
        # keyword = "Fail"
        result_label.config(text="No verdict keyword provided.")
        return
    # Parse keywords from input
    keywords = [kw.strip() for kw in keyword.split(',')]
    
    data = []
    sr_no = 1

    for folder_name in os.listdir(input_dir):
        folder_path = os.path.join(input_dir, folder_name)
        if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
            pdf_file = None
            vtest_file = None
            detailed_report_file = None
            CAN_log = None

            for file_name in os.listdir(folder_path):
                if file_name.endswith(".pdf") and file_name.startswith("Report_"):
                    pdf_file = os.path.join(folder_path, file_name)
                elif file_name.endswith(".vtestreport") and file_name.startswith("Report_"):
                    vtest_file = os.path.join(folder_path, file_name)
                elif file_name.startswith("DetailedReport_") and file_name.endswith(".pdf"):
                    detailed_report_file = os.path.join(folder_path, file_name)
                #elif file_name.startswith("Logging") and file_name.endswith(".blf"):
                elif file_name.endswith(".blf"):
                    CAN_log = os.path.join(folder_path, file_name)    

            test_pattern_name = '_'.join(folder_name.split('_')[:-2])
            timestamp = folder_name.split('_')[-2] + "_" + folder_name.split('_')[-1]
            report_name = f"Report_{timestamp}"

            if pdf_file:
                duration = extract_duration_from_pdf(pdf_file)
                sentences_with_keywords = {}
                for keyword in keywords:
                    sentences = extract_sentences_with_keywords(pdf_file, [keyword])
                    if sentences:
                        sentences_with_keywords[keyword] = sentences
                if not sentences_with_keywords:
                    results = "NA"
                else:
                    results = ""
                    for keyword, sentences in sentences_with_keywords.items():
                        if sentences:
                            results += f"{keyword} comments:\n" + "\n".join(sentences) + "\n\n"
                    results = results.strip() 
                date_from_pdf = extract_date_from_pdf(pdf_file)
                download_time, activation_time, installation_time, down_time = extract_times_from_pdf(pdf_file)
                campaign_type, program_before, program_after, rssi, setup = extract_campgtyp_prgID_rssi_setup_details_from_pdf(pdf_file)
                campaign_name, campaign_url = extract_campaign_details_from_pdf(pdf_file)
                ecu_type = determine_ecu_type(test_pattern_name)
                start_time, end_time = extract_Testtimes_from_pdf(pdf_file)
                
            else:
                duration = None
                results = "PDF file not found"
                date_from_pdf = None
                download_time = activation_time = installation_time = down_time = campaign_name = campaign_url = campaign_type = program_before = program_after = rssi = setup = ecu_type = start_time = end_time = None

            vt_report_status = "Available" if vtest_file else "Not Available"

            k_available = "Available" if detailed_report_file else "Not Available"
            k_result = extract_result_from_detailed_report(detailed_report_file) if detailed_report_file else "Not Available"
            test_case_result = extract_test_case_result_from_vtreport(pdf_file) if pdf_file else "Not Available"
            can_log = "Available" if CAN_log else "Not Available"
            
            data.append({
                        'Sr No.': sr_no,        # col A
                        'Test_Pattern_Name': test_pattern_name,  # col B
                        'Report_Name': report_name,   # col C
                        'Executed_Time': '', # col D
                        'Execution_Date': date_from_pdf, # col E
                        'Start time' : start_time,
                        'End time' : end_time,
                        'Execution_Time': duration,   # col F
                        'Download_Time': download_time,  # col G
                        'Installation_Time': installation_time,  # col H
                        'Activation_Time': activation_time,  # col I
                        'Down_Time': down_time, # col J
                        'Campaign_URL': campaign_url,  # col K
                        'Campaign_Name': campaign_name,  # col L
                        'Campaign_Type': campaign_type, # col M
                        'Program_ID_Before_Test': program_before, # col N
                        'Program_ID_After_Test': program_after, # col O
                        'IVI_RSSI_Min': '', # col P
                        'IVI_RSSI_Max': '', # col Q
                        'IVI_RSSI_Avg': rssi, # col R
                        'TCU_RSSI_Min': '', # col S
                        'TCU_RSSI_Max': '', # col T
                        'TCU_RSSI_Avg': '', # col U
                        'Cloud_Connectivity_Establishment_Time': '', # col V
                        'ECU_ROM_Type': ecu_type, # col W
                        'KITE_Report': k_available,  # col X
                        'KITE_Result': k_result,  # col Y
                        'VT_Report': vt_report_status,  # col Z
                        'VT_Result': test_case_result, # col AA
                        'Stage_Information': 'Stage 0', # col AB
                        'Failure_Type': 'Not started', # col AC
                        'Setup': setup, # col AD
                        'CAN_log' : can_log , # col AE
                        'Comments': results,  # col AF
                        'Responsible': '', # col AG
                        'RCA': '' # col AH
                        
            })
            sr_no += 1
            
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file_name = os.path.join(output_dir, f'Test_Summary_{current_time}.xlsx')

    write_sentences_to_excel(data, excel_file_name)

    #result_label.config(text=f"Excel file generated: {excel_file_name}")
    messagebox.showinfo("Excel Generation", f"Excel file generated: {excel_file_name}")
    ok_button.config(state=tk.NORMAL)

def close_app():
    excel_dir = os.path.dirname(output_entry.get())
    os.startfile(excel_dir)
    root.destroy()
    
# def find_report_viewer():
    # # Search for ReportViewerCli.exe in Program Files under Vector CANoe Test Report Viewer folders
    # program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    # for folder_name in os.listdir(program_files):
        # if folder_name.startswith("Vector CANoe Test Report Viewer"):  # Check for folders starting with "Vector CANoe Test Report Viewer"
            # vector_dir = os.path.join(program_files, folder_name)
            # for root, dirs, files in os.walk(vector_dir):
                # if "ReportViewerCli.exe" in files:
                    # return os.path.join(root, "ReportViewerCli.exe")
    # return None
def find_report_viewers():
    # Search for all ReportViewerCli.exe paths in Program Files under Vector CANoe Test Report Viewer folders
    program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    viewer_paths = []
    for folder_name in os.listdir(program_files):
        if folder_name.startswith("Vector CANoe Test Report Viewer"):  # Check for folders starting with "Vector CANoe Test Report Viewer"
            vector_dir = os.path.join(program_files, folder_name)
            for root, dirs, files in os.walk(vector_dir):
                if "ReportViewerCli.exe" in files:
                    viewer_paths.append(os.path.join(root, "ReportViewerCli.exe"))    
    return viewer_paths 

def create_report_pdfs():
    input_dir = input_entry.get()
    if not input_dir:
        result_label.config(text="Please select input folder.")
        return

    report_viewer_paths = find_report_viewers()
    if not report_viewer_paths:
        messagebox.showerror("Error", "No compatible ReportViewerCli.exe found.")
        return

    try:
        for folder_name in os.listdir(input_dir):
            folder_path = os.path.join(input_dir, folder_name)
            if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
                for file_name in os.listdir(folder_path):
                    if file_name.endswith(".vtestreport") and file_name.startswith("Report_"):
                        vtestreport_path = os.path.join(folder_path, file_name)
                        pdf_path = vtestreport_path.replace(".vtestreport", ".pdf")

                        if os.path.exists(pdf_path):
                            print(f"Skipping {vtestreport_path}, PDF already exists: {pdf_path}")
                            continue

                        success = False
                        for report_viewer_path in report_viewer_paths:
                            try:
                                print(f"Trying {report_viewer_path} for {vtestreport_path}")
                                subprocess.run(
                                    [
                                        report_viewer_path,
                                        f"-r={vtestreport_path}",
                                        f"-p={pdf_path}",
                                        "-s=A4",
                                        "-o=Portrait",
                                        #"-nc",
                                        #"-nr",
                                        "-ne",
                                    ],
                                    check=True,
                                    capture_output=True,
                                    text=True,
                                )
                                print(f"Success with {report_viewer_path}")
                                success = True
                                break
                            except subprocess.CalledProcessError as e:
                                error_message = e.stderr or str(e)
                                print(f"Failed with {report_viewer_path}: {error_message}")
                                if "IncompatibleReportException" in error_message:
                                    continue  # Try next executable
                                elif e.returncode:  # Handle other return codes flexibly
                                    continue  # Log and move to the next version
                                else:
                                    raise  # Stop for unexpected errors

                        if not success:
                            print(f"No compatible ReportViewerCli.exe found for {vtestreport_path}")
                            raise Exception(f"No compatible ReportViewerCli.exe found for {vtestreport_path}")

        messagebox.showinfo("PDF Creation", "PDF reports created successfully.")
    except Exception as e:
        print(f"Exception encountered: {str(e)}")
        messagebox.showerror("PDF Creation Error", f"Error creating PDF reports: {str(e)}")

    ok_button.config(state=tk.NORMAL)


# def create_report_pdfs():
    # input_dir = input_entry.get()
    # if not input_dir:
        # result_label.config(text="Please select input folder.")
        # return
        
    # report_viewer_path = find_report_viewer()
    # if not report_viewer_path:
        # messagebox.showerror("Error", "ReportViewerCli.exe not found in Vector CANoe Test Report Viewer.")
        # return


    # try:
        # for folder_name in os.listdir(input_dir):
            # folder_path = os.path.join(input_dir, folder_name)
            # if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
                # for file_name in os.listdir(folder_path):
                    # if file_name.endswith(".vtestreport") and file_name.startswith("Report_"):
                        # vtestreport_path = os.path.join(folder_path, file_name)
                        # pdf_path = vtestreport_path.replace(".vtestreport", ".pdf")
                        # subprocess.run([
                            # report_viewer_path, #dynamic path
                            # f"-r={vtestreport_path}",
                            # f"-p={pdf_path}",
                            # "-s=A4",
                            # "-o=Portrait",
                            # "-nc",
                            # "-nr",
                            # "-ne"
                        # ])
                        # #result_label.config(text="PDF reports created successfully.")
        # messagebox.showinfo("PDF Creation", "PDF reports created successfully.")
    # except Exception as e:
        # messagebox.showerror("PDF Creation Error", f"Error creating PDF reports: {str(e)}")

    # ok_button.config(state=tk.NORMAL)

def delete_report_pdfs():
    input_dir = input_entry.get()

    if not input_dir:
        result_label.config(text="Please select the input folder.")
        return

    confirm_delete = tk.messagebox.askyesno("Confirm Delete", "Are you sure you want to delete the report PDFs in the selected date folder?")

    if not confirm_delete:
        return

    deleted_files = []
    errors = []  # Initialize empty list for errors
    for folder_name in os.listdir(input_dir):
        folder_path = os.path.join(input_dir, folder_name)
        if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
            for file_name in os.listdir(folder_path):
                if file_name.endswith(".pdf") and file_name.startswith("Report_"):
                    pdf_path = os.path.join(folder_path, file_name)
                    try:
                        # Check if the file exists before attempting to delete it
                        if os.path.exists(pdf_path):
                            send2trash.send2trash(os.path.abspath(pdf_path))
                            deleted_files.append(pdf_path)
                        else:
                            print(f"File not found: {pdf_path}")
                    except Exception as e:
                        print(f"Error deleting file {pdf_path}: {str(e)}")

    # if deleted_files:
        # #result_label.config(text=f"Deleted files:\n" + "\n".join(deleted_files))
        # messagebox.showerror("Deletion Errors", "\n".join(errors))
    # else:
        # #result_label.config(text="No report PDFs found to delete.")
        # messagebox.showinfo("Deletion Complete", "PDF files deleted successfully.")
    if deleted_files:
        messagebox.showinfo("Deletion Complete", "PDF files deleted successfully.")
    else:
        messagebox.showinfo("Deletion Complete", "No report PDFs found to delete.")    
    ok_button.config(state=tk.NORMAL)    
    
    
def reset_fields():
    input_entry.delete(0, tk.END)
    output_entry.delete(0, tk.END)
    #keyword_entry.delete(0, tk.END)
    result_label.config(text="")
    
def create_delivery_folder():
    input_dir = input_entry.get()  # Get folder path from GUI input

    if not input_dir:
        result_label.config(text="Please select the input folder.")
        return

    confirm_delete = tk.messagebox.askyesno(
        "Confirm Cleanup",
        "Are you sure you want to delete unwanted files in all the selected TS_ folders?"
    )

    if not confirm_delete:
        return

    deleted_files = []
    for folder_name in os.listdir(input_dir):
        folder_path = os.path.join(input_dir, folder_name)

        # Process only folders that start with "TS_"
        if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
            for file_name in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file_name)

                # Skip directories
                if os.path.isdir(file_path):
                    continue  

                # Check if the file matches the keep criteria
                if (file_name.endswith(".blf") or 
                    file_name.endswith(".vtestreport") or 
                    (file_name.endswith(".pdf") and file_name.startswith("DetailedReport_"))):
                    continue  # Keep these files
                
                # Delete unwanted files
                try:
                    if os.path.exists(file_path):
                        send2trash.send2trash(os.path.abspath(file_path))  # Move to trash
                        deleted_files.append(file_path)
                    else:
                        print(f"File not found: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {str(e)}")

    # Show result message
    if deleted_files:
        messagebox.showinfo("Cleanup Complete", "Unwanted files have been moved to trash.")
    else:
        messagebox.showinfo("Cleanup Complete", "No unwanted files found for deletion.")
    
    ok_button.config(state=tk.NORMAL)
    
    
# def create_delivery_folder():
    # input_dir = input_entry.get()

    # if not input_dir:
        # result_label.config(text="Please select the input folder.")
        # return

    # confirm_delete = tk.messagebox.askyesno("Confirm Delete", "Are you sure you want to delete the files in the selected date folder?")

    # if not confirm_delete:
        # return

    # deleted_files = []
    # errors = []  # Initialize empty list for errors
    # for folder_name in os.listdir(input_dir):
        # folder_path = os.path.join(input_dir, folder_name)
        # if os.path.isdir(folder_path) and folder_name.startswith("TS_"):
            # for file_name in os.listdir(folder_path):
                # if file_name.endswith(".pdf") and file_name.startswith("Detailed"):
                    # pdf_path = os.path.join(folder_path, file_name)
                    # try:
                        # # Check if the file exists before attempting to delete it
                        # if os.path.exists(pdf_path):
                            # send2trash.send2trash(os.path.abspath(pdf_path))
                            # deleted_files.append(pdf_path)
                        # else:
                            # print(f"File not found: {pdf_path}")
                    # except Exception as e:
                        # print(f"Error deleting file {pdf_path}: {str(e)}")

    # # if deleted_files:
        # # #result_label.config(text=f"Deleted files:\n" + "\n".join(deleted_files))
        # # messagebox.showerror("Deletion Errors", "\n".join(errors))
    # # else:
        # # #result_label.config(text="No report PDFs found to delete.")
        # # messagebox.showinfo("Deletion Complete", "PDF files deleted successfully.")
    # if deleted_files:
        # messagebox.showinfo("Deletion Complete", "PDF files deleted successfully.")
    # else:
        # messagebox.showinfo("Deletion Complete", "No report PDFs found to delete.")    
    # ok_button.config(state=tk.NORMAL)   

# Create the main window
root = tk.Tk()
root.title("Test Summary Tool v3.9")
  

# Label and Entry for input folder
input_label = tk.Label(root, text="Input Folder (Date folder):")
input_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
input_entry = tk.Entry(root, width=50)
input_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
browse_input_button = tk.Button(root, text="Browse", command=browse_input_folder)
browse_input_button.grid(row=0, column=2, padx=5, pady=5)

# Label and Entry for output folder
output_label = tk.Label(root, text="Output Folder:")
output_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
browse_output_button = tk.Button(root, text="Browse", command=browse_output_dir)
browse_output_button.grid(row=1, column=2, padx=5, pady=5)

# Label and Entry for keyword
keyword_label = tk.Label(root, text="Verdict Keywords:\n(comma separated)")
keyword_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
keyword_entry = tk.Entry(root, width=50)
keyword_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
# Prefill with default keywords
default_keywords = "Fail,Inconclusive"
keyword_entry.insert(0, default_keywords)

# Generate button
generate_button = tk.Button(root, text="Generate Excel", command=generate_excel)
generate_button.grid(row=3, columnspan=2, padx=0, pady=5)

# Create report pdf button
create_pdf_button = tk.Button(root, text="Create Report PDFs", command=create_report_pdfs)
create_pdf_button.grid(row=3, column=0, padx=2, pady=5)

# Delete report pdf button
delete_pdf_button = tk.Button(root, text="Delete Report PDFs", command=delete_report_pdfs)
delete_pdf_button.grid(row=3, column=2, padx=1, pady=5)

# Delivery folder creation button
create_delivery_button = tk.Button(root, text="Create Delivery folder", command=create_delivery_folder)
create_delivery_button.grid(row=4, columnspan=2, padx=1, pady=5)

# Result label
result_label = tk.Label(root, text="")
result_label.grid(row=5, columnspan=3, padx=5, pady=5)

# OK button to close the application
ok_button = tk.Button(root, text="OK", command=close_app, state=tk.DISABLED)
ok_button.grid(row=6, columnspan=3, padx=5, pady=5)

# Reset Button 
reset_button = tk.Button(root, text="Clear Paths", command=reset_fields)
reset_button.grid(row=2, column=2, padx=5, pady=5)

root.mainloop()
